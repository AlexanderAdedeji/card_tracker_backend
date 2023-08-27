import os
from typing import List
from fastapi import APIRouter, Depends
from sqlalchemy.orm import Session
from starlette.status import (
    HTTP_500_INTERNAL_SERVER_ERROR,
    HTTP_201_CREATED,
    HTTP_200_OK,
)
from openpyxl import load_workbook
from fastapi import Depends, HTTPException, APIRouter
from app.commonLib.schemas.response_model import ResponseWrapper
from app.models.collection_centers_models import CollectionCentre
from app.models.local_goverment_models import LocalGovernment
from app.repositories.location_repo import local_govt_repo, collection_center_repo

from app.api.dependencies.db import get_db
from app.models.local_goverment_models import LocalGovernment
from app.schemas.location_schemas import LocalGovernmentResponse

router = APIRouter()


@router.get("/location")
def location():
    return {"message": "Location"}


router = APIRouter()


@router.post("/populate_local_governments", status_code=HTTP_201_CREATED)
def populate_local_government(db=Depends(get_db)):
    try:
        location = os.getcwd()

        wb = load_workbook(f"{location}/app/excel_files/local_govts.xlsx")
        ws = wb.active
        local_government = ws["B"]

        batch_size = 100  # Adjust as needed
        batch = []

        for i in range(2, len(local_government) + 1):
            name = ws[f"B{i}"].value
            code = ws[f"D{i}"].value
            db_local_govt = LocalGovernment(name=name, code=code)
            batch.append(db_local_govt)

            if len(batch) >= batch_size:
                local_govt_repo.create(db, obj_in=db_local_govt)

                batch = []

        if batch:
            db.add_all(batch)
            db.commit()

        return {"message": "Local Governments Created Successfully"}

    except Exception as e:
        return HTTPException(status_code=HTTP_500_INTERNAL_SERVER_ERROR, detail=str(e))


@router.post("/populate_collection_centres", status_code=HTTP_201_CREATED)
def populate_collection_centres(db: Session = Depends(get_db)):
    try:
        location = os.getcwd()
        wb = load_workbook(f"{location}/app/excel_files/CollectionCentres.xlsx")
        ws = wb.active
        collection_centres = ws["D"]

        batch_size = 100  # Adjust as needed
        batch = []

        for i in range(2, len(collection_centres) + 1):
            name = ws[f"D{i}"].value
            code = ws[f"B{i}"].value
            localGovtCode = ws[f"C{i}"].value
            db_collection_centres = CollectionCentre(
                name=name, code=code, local_govt_code=localGovtCode
            )
            batch.append(db_collection_centres)

            if len(batch) >= batch_size:
                collection_center_repo.create(db, obj_in=db_collection_centres)
                batch = []

        if batch:
            db.add_all(batch)
            db.commit()

        return {"message": "Collection Centres Created Successfully"}

    except Exception as e:
        return HTTPException(status_code=500, detail=str(e))


@router.get(
    "/collection_centers",
    status_code=HTTP_200_OK
    #   response_model=ResponseWrapper
)
def all_collection_centers(db: Session = Depends(get_db)):
    collection_centers = collection_center_repo.get_all(db)
    return collection_centers
    # ResponseWrapper(status_code=HTTP_200_OK, message="All Collection centers")


@router.get(
    "/local_governments",
    status_code=HTTP_200_OK,
    # response_model=LocalGovernmentResponse,
)
def all_local_governments(db: Session = Depends(get_db)):
    local_governments = local_govt_repo.get_all(db)
    return local_governments


@router.get(
    "/get_local_government/{lga_code}"
    # ,
    # response_model=List[LocalGovt]
)
def get_local_government(lag_code: str, db: Session = Depends(get_db)):
    local_governments = local_govt_repo.get_by_field(
        db, field_name="code", field_value=lag_code
    )
    return local_governments


@router.get(
    "/get_local_government_collection_centers/{lga_code}"
    # ,
    # response_model=List[LocalGovt]
)
def get_collection_centers_in_local_government(
    lag_code: str, db: Session = Depends(get_db)
):
    collection_centers = collection_center_repo.get_by_field(
        db, field_name="local_govt_code", field_value=lag_code
    )
    return collection_centers
